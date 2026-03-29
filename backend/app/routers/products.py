import json
import re
import uuid
from io import BytesIO

import openpyxl
from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, status
from sqlalchemy.orm import Session

from app.core.database import get_db
from app.dependencies import get_current_seller
from app.models.product import Product, ProductVariant
from app.models.seller import Seller
from app.schemas.product import ProductCreate, ProductOut, ProductUpdate

router = APIRouter(prefix="/api/products", tags=["products"])


def _slugify(text: str) -> str:
    text = text.lower().strip()
    text = re.sub(r'[^\w\s-]', '', text)
    text = re.sub(r'[\s_]+', '-', text)
    text = re.sub(r'-+', '-', text)
    return text.strip('-') or 'product'


def _unique_slug(db: Session, seller_id, name: str, exclude_id=None) -> str:
    base = _slugify(name)
    slug = base
    counter = 2
    while True:
        q = db.query(Product).filter(
            Product.seller_id == seller_id,
            Product.slug == slug,
        )
        if exclude_id:
            q = q.filter(Product.id != exclude_id)
        if not q.first():
            return slug
        slug = f"{base}-{counter}"
        counter += 1


def _sync_variants(db: Session, product: Product, variants_data: list) -> None:
    """Replace all variants for a product with new data."""
    for v in product.variants:
        db.delete(v)
    db.flush()

    for i, v in enumerate(variants_data):
        db.add(ProductVariant(
            product_id=product.id,
            label=v.label,
            price=v.price,
            is_available=v.is_available,
            sort_order=i,
        ))


def _set_base_price(product: Product, variants_data: list) -> None:
    """When variants exist, set base price to the minimum variant price."""
    if variants_data:
        product.price = min(v.price for v in variants_data)


@router.get("", response_model=list[ProductOut])
def list_products(
    db: Session = Depends(get_db),
    current_seller: Seller = Depends(get_current_seller),
):
    return (
        db.query(Product)
        .filter(Product.seller_id == current_seller.id)
        .order_by(Product.created_at.desc())
        .all()
    )


@router.post("", response_model=ProductOut, status_code=status.HTTP_201_CREATED)
def create_product(
    payload: ProductCreate,
    db: Session = Depends(get_db),
    current_seller: Seller = Depends(get_current_seller),
):
    data = payload.model_dump(exclude={"variants", "image_urls"})
    data["image_urls"] = json.dumps(payload.image_urls)
    if payload.image_urls:
        data["image_url"] = payload.image_urls[0]
    data["slug"] = _unique_slug(db, current_seller.id, payload.name)
    product = Product(**data, seller_id=current_seller.id)

    if payload.has_variants and payload.variants:
        _set_base_price(product, payload.variants)

    db.add(product)
    db.flush()

    if payload.has_variants and payload.variants:
        for i, v in enumerate(payload.variants):
            db.add(ProductVariant(
                product_id=product.id,
                label=v.label,
                price=v.price,
                is_available=v.is_available,
                sort_order=i,
            ))

    db.commit()
    db.refresh(product)
    return product


@router.get("/{product_id}", response_model=ProductOut)
def get_product(
    product_id: uuid.UUID,
    db: Session = Depends(get_db),
    current_seller: Seller = Depends(get_current_seller),
):
    product = (
        db.query(Product)
        .filter(Product.id == product_id, Product.seller_id == current_seller.id)
        .first()
    )
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    return product


@router.put("/{product_id}", response_model=ProductOut)
def update_product(
    product_id: uuid.UUID,
    payload: ProductUpdate,
    db: Session = Depends(get_db),
    current_seller: Seller = Depends(get_current_seller),
):
    product = (
        db.query(Product)
        .filter(Product.id == product_id, Product.seller_id == current_seller.id)
        .first()
    )
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")

    update_data = payload.model_dump(exclude_unset=True, exclude={"variants", "image_urls"})
    for field, value in update_data.items():
        setattr(product, field, value)

    if payload.name is not None:
        product.slug = _unique_slug(db, current_seller.id, payload.name, exclude_id=product_id)

    if payload.image_urls is not None:
        product.image_urls = json.dumps(payload.image_urls)
        product.image_url = payload.image_urls[0] if payload.image_urls else None

    if payload.variants is not None:
        _sync_variants(db, product, payload.variants)
        if product.has_variants and payload.variants:
            _set_base_price(product, payload.variants)

    db.commit()
    db.refresh(product)
    return product


@router.delete("/{product_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_product(
    product_id: uuid.UUID,
    db: Session = Depends(get_db),
    current_seller: Seller = Depends(get_current_seller),
):
    product = (
        db.query(Product)
        .filter(Product.id == product_id, Product.seller_id == current_seller.id)
        .first()
    )
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")

    db.delete(product)
    db.commit()


@router.post("/bulk-import")
async def bulk_import_products(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_seller: Seller = Depends(get_current_seller),
):
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx files are supported.")

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
        ws = wb.active
    except Exception:
        raise HTTPException(status_code=400, detail="Could not read the Excel file. Make sure it is a valid .xlsx file.")

    headers = [str(cell.value).strip().lower() if cell.value else "" for cell in ws[1]]
    required = {"name", "price"}
    if not required.issubset(set(headers)):
        raise HTTPException(status_code=400, detail="Missing required columns: 'name' and 'price' must be present.")

    def col(row, name):
        try:
            idx = headers.index(name)
            val = row[idx].value
            return str(val).strip() if val is not None else ""
        except ValueError:
            return ""

    # Group rows by product name (case-insensitive, trimmed)
    groups: dict[str, list] = {}
    errors: list[str] = []

    for row_num, row in enumerate(ws.iter_rows(min_row=2), start=2):
        if all(cell.value is None for cell in row):
            continue

        name = col(row, "name").strip()
        if not name:
            # Continuation row for a variant — attach to last group
            if groups:
                last_key = list(groups.keys())[-1]
                groups[last_key].append((row_num, row))
            continue

        key = name.lower()
        if key not in groups:
            groups[key] = []
        groups[key].append((row_num, row))

    created = 0
    failed = 0

    for key, rows in groups.items():
        first_row_num, first_row = rows[0]
        name = col(first_row, "name").strip() or key

        # Determine if any row has a variant_label
        has_variant_col = "variant_label" in headers
        variant_rows = []
        if has_variant_col:
            for rn, r in rows:
                vl = col(r, "variant_label")
                vp = col(r, "variant_price") if "variant_price" in headers else ""
                if vl:
                    variant_rows.append((rn, vl, vp))

        has_variants = len(variant_rows) > 0

        # Validate base price (only needed when no variants)
        base_price = None
        if not has_variants:
            raw_price = col(first_row, "price")
            try:
                base_price = float(raw_price)
            except (ValueError, TypeError):
                errors.append(f"Row {first_row_num}: '{name}' — price is missing or invalid.")
                failed += 1
                continue

        # Validate variant prices
        if has_variants:
            valid_variants = []
            variant_error = False
            for rn, vl, vp in variant_rows:
                try:
                    valid_variants.append((vl, float(vp)))
                except (ValueError, TypeError):
                    errors.append(f"Row {rn}: variant '{vl}' of '{name}' — variant_price is missing or invalid.")
                    variant_error = True
            if variant_error:
                failed += 1
                continue

        # Parse optional fields from first row
        raw_cap = col(first_row, "compare_at_price")
        try:
            compare_at_price = float(raw_cap) if raw_cap else None
        except (ValueError, TypeError):
            compare_at_price = None

        is_available_raw = col(first_row, "is_available").upper()
        is_available = is_available_raw != "FALSE"

        description = col(first_row, "description") or None
        image_url = col(first_row, "image_url") or None
        image_urls = json.dumps([image_url] if image_url else [])

        # Create product
        product = Product(
            seller_id=current_seller.id,
            name=name,
            slug=_unique_slug(db, current_seller.id, name),
            description=description,
            price=base_price if not has_variants else min(p for _, p in valid_variants),
            compare_at_price=compare_at_price,
            image_url=image_url,
            image_urls=image_urls,
            is_available=is_available,
            has_variants=has_variants,
        )
        db.add(product)
        db.flush()

        if has_variants:
            for i, (vl, vp) in enumerate(valid_variants):
                db.add(ProductVariant(
                    product_id=product.id,
                    label=vl,
                    price=vp,
                    is_available=True,
                    sort_order=i,
                ))

        created += 1

    db.commit()

    return {"created": created, "failed": failed, "errors": errors}
